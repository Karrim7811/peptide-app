"""
Regenerates src/lib/peptide-knowledge.ts from the Excel spreadsheet.
Run with: python generate_knowledge.py
"""
import openpyxl
import re

def clean(s):
    s = str(s or '').strip()
    s = s.replace("'", "\\'").replace('`', '\\`').replace('\n', ' ').replace('\r', ' ')
    return s

def int_cv(v):
    try:
        return int(float(str(v)))
    except Exception:
        return 0

wb = openpyxl.load_workbook(
    'Peptides_Master_List_FULL_Explainers_CV_Interactions_Dropdowns.xlsx',
    read_only=True
)

# --- Master sheet ---
ws1 = wb['Master Peptide List']
master_rows = list(ws1.iter_rows(values_only=True))
mh = master_rows[0]

# --- Peptide Matrix sheet ---
ws2 = wb['Peptide Matrix']
matrix_rows = list(ws2.iter_rows(values_only=True))
matrix_headers = list(matrix_rows[0])
cats = matrix_headers[1:]  # skip 'Peptide'

# Build category lookup {name -> [cat, ...]}
matrix = {}
for row in matrix_rows[1:]:
    if not row[0]:
        continue
    name = row[0]
    active_cats = [cats[i] for i, v in enumerate(row[1:]) if v == '\u2714']
    matrix[name] = active_cats

# Build master lookup
master = {}
for row in master_rows[1:]:
    if not row[0]:
        continue
    d = dict(zip(mh, row))
    master[d['Peptide Name']] = d

# Compute stacksWellWith: shares a category but different primary class
def get_stacks(name, my_cats, my_goal_cat):
    """
    Returns peptides from DIFFERENT goal categories — complementary stacks.
    E.g. a Metabolic peptide stacks well with GH Axis, Healing, Muscle peptides.
    Excludes same primary goal category entirely.
    """
    results = []
    my_goal = (my_goal_cat or '').strip().lower()
    for other_name, other_cats in matrix.items():
        if other_name.strip() == name.strip():
            continue
        other_d = master.get(other_name, {})
        other_goal = str(other_d.get('Goal Category (standardized)', '') or '').strip().lower()
        # Must be a different goal category
        if my_goal and other_goal and my_goal == other_goal:
            continue
        # Must have at least one meaningful category (not empty)
        if not other_cats:
            continue
        results.append(other_name)
    # Sort by: prefer peptides with overlapping benefit areas first, then alphabetical
    def score(n):
        o_cats = matrix.get(n, [])
        shared = len(set(my_cats) & set(o_cats))
        return -shared  # more shared = lower score = first
    results.sort(key=score)
    return results[:10]

lines = []
lines.append('// Auto-generated from Peptides_Master_List_FULL_Explainers_CV_Interactions_Dropdowns.xlsx')
lines.append('// Run: python generate_knowledge.py  to regenerate')
lines.append('')
lines.append('export interface PeptideKnowledge {')
lines.append('  name: string')
lines.append('  primaryPurpose: string')
lines.append('  whatItDoes: string')
lines.append('  commonUseExamples: string')
lines.append('  dosageRange: string')
lines.append('  riskCautions: string')
lines.append('  bestFor: string')
lines.append('  keyEffects: string')
lines.append('  evidenceLevel: string')
lines.append('  bottomLine: string')
lines.append('  avoidIf: string')
lines.append('  cvRating: number')
lines.append('  cvNotes: string')
lines.append('  drugInteractions: string')
lines.append('  goalCategory: string')
lines.append('  goalCategories: string[]')
lines.append('  stacksWellWith: string[]')
lines.append('}')
lines.append('')
lines.append('export const PEPTIDE_KNOWLEDGE: PeptideKnowledge[] = [')

for row in master_rows[1:]:
    if not row[0]:
        continue
    d = dict(zip(mh, row))
    n = d['Peptide Name']
    my_cats = matrix.get(n, [])
    stacks = get_stacks(n, my_cats, d.get('Goal Category (standardized)', ''))
    cv_num = int_cv(d.get('Cardiovascular Evidence/Impact Rating (0-5)', 0))

    cats_js = '[' + ', '.join(f"'{c}'" for c in my_cats) + ']'
    stacks_js = '[' + ', '.join(f"'{clean(s)}'" for s in stacks) + ']'

    lines.append('  {')
    lines.append(f"    name: '{clean(n)}',")
    lines.append(f"    primaryPurpose: '{clean(d.get('Primary Purpose', ''))}',")
    lines.append(f"    whatItDoes: '{clean(d.get('What It Does', ''))}',")
    lines.append(f"    commonUseExamples: '{clean(d.get('Common Use Examples', ''))}',")
    lines.append(f"    dosageRange: '{clean(d.get('Dosage Range (if established)', ''))}',")
    lines.append(f"    riskCautions: '{clean(d.get('Risk / Cautions', ''))}',")
    lines.append(f"    bestFor: '{clean(d.get('Best For (Practical Goal)', ''))}',")
    lines.append(f"    keyEffects: '{clean(d.get('Key Effects (Plain English)', ''))}',")
    lines.append(f"    evidenceLevel: '{clean(d.get('Evidence Level', ''))}',")
    lines.append(f"    bottomLine: '{clean(d.get('Bottom Line (1 sentence)', ''))}',")
    lines.append(f"    avoidIf: '{clean(d.get('Avoid / Not Ideal If', ''))}',")
    lines.append(f'    cvRating: {cv_num},')
    lines.append(f"    cvNotes: '{clean(d.get('CV Notes (why this rating)', ''))}',")
    lines.append(f"    drugInteractions: '{clean(d.get('Notable Drug Interactions / Monitoring', ''))}',")
    lines.append(f"    goalCategory: '{clean(d.get('Goal Category (standardized)', ''))}',")
    lines.append(f'    goalCategories: {cats_js},')
    lines.append(f'    stacksWellWith: {stacks_js},')
    lines.append('  },')

lines.append(']')
lines.append('')
lines.append('export const PEPTIDE_CATEGORIES = [')
for c in cats:
    lines.append(f"  '{c}',")
lines.append(']')
lines.append('')
# Legacy alias used by reference page
lines.append('export const GOAL_CATEGORIES = PEPTIDE_CATEGORIES')
lines.append('')
# Helper used by check-interaction route
lines.append('export function findPeptide(name: string): PeptideKnowledge | undefined {')
lines.append('  const n = name.toLowerCase()')
lines.append('  return PEPTIDE_KNOWLEDGE.find(p => p.name.toLowerCase().includes(n))')
lines.append('}')
lines.append('')
lines.append('export function searchPeptides(query: string): PeptideKnowledge[] {')
lines.append('  if (!query.trim()) return PEPTIDE_KNOWLEDGE')
lines.append('  const q = query.toLowerCase()')
lines.append('  return PEPTIDE_KNOWLEDGE.filter(p =>')
lines.append('    p.name.toLowerCase().includes(q) ||')
lines.append('    p.whatItDoes.toLowerCase().includes(q) ||')
lines.append('    p.keyEffects.toLowerCase().includes(q) ||')
lines.append('    p.goalCategory.toLowerCase().includes(q)')
lines.append('  )')
lines.append('}')
lines.append('')
lines.append('export function getPeptidesByCategory(category: string): PeptideKnowledge[] {')
lines.append('  return PEPTIDE_KNOWLEDGE.filter(p => p.goalCategories.includes(category))')
lines.append('}')
lines.append('')

output = '\n'.join(lines)
with open('src/lib/peptide-knowledge.ts', 'w', encoding='utf-8') as f:
    f.write(output)

print(f'Done. {len(master)} peptides written with goalCategories + stacksWellWith.')
